import uuid
from typing import Any
from typing import List

from fastapi import (
    APIRouter,
    HTTPException,
    BackgroundTasks,
    UploadFile,
    File,
    Query,
    Body,
    Depends,
)
from fastapi.responses import FileResponse
from sqlmodel import func, select, or_, and_, delete
from sqlalchemy.orm import aliased, selectinload
from sqlalchemy.exc import IntegrityError
from openpyxl import Workbook, load_workbook
import csv
from datetime import date
import tempfile

from app.api.deps import CurrentUser, SessionDep

from app.models import (
    Switch,
    SwitchConfigBase,
    SwitchCreate,
    SwitchPublic,
    SwitchesPublic,
    SwitchUpdate,
    Message,
    Branch,
    Inspector,
    SwitchLoginType,
    SwitchConfig,
    SwitchConfigShow,
    SwitchLite,
)

from app.services.switches.config.backup import download_config

router = APIRouter(prefix="/switches", tags=["switches"])


def base_query_method(
    search_text: str = "",
    branch_ids: list[str] = Query(default=[], alias="branch_ids"),
    status: list[str] = Query(default=[], alias="status"),
):
    today = date.today()
    # 搜索条件
    search_param = (
        or_(
            Switch.name.ilike(f"%{search_text}%"),
            Switch.ip.ilike(f"%{search_text}%"),
            Switch.hardware_type.ilike(f"%{search_text}%"),
            Switch.description.ilike(f"%{search_text}%"),
        )
        if search_text
        else True
    )

    # 组合搜索和分支过滤条件
    filters = [search_param]

    if branch_ids:
        filters.append(Switch.branch_id.in_(branch_ids))

    combined_filter = and_(*filters)

    # join with latest_config and filter by status
    LatestConfig = aliased(SwitchConfig)
    base_query = select(Switch, LatestConfig)

    latest_subq = (
        select(
            SwitchConfig.switch_id,
            func.max(SwitchConfig.created).label("max_created"),
        )
        .group_by(SwitchConfig.switch_id)
        .subquery()
    )

    base_query = base_query.outerjoin(
        latest_subq, Switch.id == latest_subq.c.switch_id
    ).outerjoin(
        LatestConfig,
        (LatestConfig.switch_id == latest_subq.c.switch_id)
        & (LatestConfig.created == today),
    )

    if status:
        if len(status) > 1:
            pass
        else:
            status = status[0]
            if status == "success":
                base_query = base_query.where(LatestConfig.status == "success")
            else:
                base_query = base_query.where(LatestConfig.status != "success")

    base_query = base_query.where(combined_filter)

    # sorted by id
    base_query = base_query.order_by(Switch.ip)

    return base_query


@router.get("/", response_model=SwitchesPublic)
def read_switches(
    session: SessionDep,
    current_user: CurrentUser,
    skip: int = 0,
    limit: int = 100,
    search_text: str = "",
    branch_ids: list[str] = Query(
        default=[], alias="branch_ids"
    ),  # 接收逗号分隔的字符串
    status: list[str] = Query(default=[], alias="status"),
) -> Any:
    """
    Retrieve switches with latest config.
    """
    base_query = base_query_method(search_text, branch_ids, status)

    # 总数
    count = session.exec(select(func.count()).select_from(base_query.subquery())).one()

    # 查询 Switch 列表 + eager load latest_config
    rows = session.exec(
        base_query.options(
            selectinload(Switch.branch),
            selectinload(Switch.login_type),
        )
        .offset(skip)
        .limit(limit)
    ).all()

    switches_data = [
        SwitchPublic(
            **switch.model_dump(),
            branch=switch.branch,
            login_type=switch.login_type,
            inspector=switch.inspector,
            latest_config=SwitchConfigBase.model_validate(latest) if latest else None,
        )
        for switch, latest in rows
    ]

    return SwitchesPublic(data=switches_data, count=count)


@router.get("/read/{id}", response_model=SwitchLite)
def read_switch(
    session: SessionDep,
    current_user: CurrentUser,
    id: uuid.UUID,
) -> Any:
    """
    Get switch by ID.
    
    Requires authentication. The current_user parameter triggers automatic
    permission validation through dependency injection.
    """
    # Permission check is handled by CurrentUser dependency
    if not current_user:
        raise HTTPException(status_code=403, detail="Not enough permissions")
    
    switch = session.get(Switch, id)
    if not switch:
        raise HTTPException(status_code=404, detail="Switch not found")
    return switch


@router.post("/", response_model=Switch)
def create_switch(
    *, session: SessionDep, current_user: CurrentUser, switch_in: SwitchCreate
) -> Any:
    """
    Create new switch.
    
    Requires authentication. The current_user parameter triggers automatic
    permission validation through dependency injection.
    """
    # Permission check is handled by CurrentUser dependency
    if not current_user:
        raise HTTPException(status_code=403, detail="Not enough permissions")
    
    ip = switch_in.ip
    # Pre-check to return a friendly 409 if IP already exists
    if session.exec(select(Switch).where(Switch.ip == ip)).first():
        raise HTTPException(
            status_code=409,
            detail="设备IP已存在",
        )

    switch = Switch.model_validate(switch_in)
    session.add(switch)
    try:
        session.commit()
        session.refresh(switch)
    except IntegrityError:
        # In case of a race condition where another process inserted the same IP
        session.rollback()
        raise HTTPException(
            status_code=409,
            detail="设备IP已存在",
        )
    return switch


@router.put("/update/{id}", response_model=SwitchLite)
def update_switch(
    *,
    session: SessionDep,
    current_user: CurrentUser,
    id: uuid.UUID,
    switch_in: SwitchUpdate,
) -> Any:
    """
    Update a switch.
    """
    switch = session.get(Switch, id)
    if not switch:
        raise HTTPException(status_code=404, detail="设备不存在")

    update_dict = switch_in.model_dump(exclude_unset=True)
    switch.sqlmodel_update(update_dict)
    session.add(switch)
    try:
        session.commit()
        session.refresh(switch)
    except IntegrityError:
        session.rollback()
        # IP uniqueness conflict (or other integrity constraint)
        raise HTTPException(
            status_code=409,
            detail="设备不存在",
        )
    return switch


@router.delete("/delete/{id}")
def delete_switch(session: SessionDep, current_user: CurrentUser, id: uuid.UUID) -> Message:
    """
    Delete a switch.
    """
    switch = session.get(Switch, id)
    if not switch:
        raise HTTPException(status_code=404, detail="设备不存在")

    session.delete(switch)
    session.commit()
    return Message(message="设备删除成功")


@router.post("/delete")
def delete_multiple_switches(
    session: SessionDep,
    current_user: CurrentUser,
    ids: List[uuid.UUID] = Body(...),
) -> Message:
    """
    Delete multiple switches.
    """
    # 删除所有匹配的设备
    session.exec(delete(Switch).where(Switch.id.in_(ids)))
    session.commit()

    return Message(message="设备删除成功")


@router.get("/backup/{id}")
def backup_switch(
    session: SessionDep, background_tasks: BackgroundTasks, current_user: CurrentUser, id: uuid.UUID
) -> Any:
    """
    Backup a switch.
    """
    switch = session.get(Switch, id)
    if not switch:
        raise HTTPException(status_code=404, detail="设备不存在")

    # Implement your backup logic here
    # For example, you might copy the switch data to a backup table
    background_tasks.add_task(download_config, switch.id)
    return Message(message="Switch backed up successfully")


@router.post("/backup")
def backup_multiple_switches(
    session: SessionDep,
    background_tasks: BackgroundTasks,
    current_user: CurrentUser,
    ids: List[uuid.UUID] = Body(...),
) -> Any:
    """
    Backup multiple switches.
    """
    # 查询所有匹配的设备
    switches = session.exec(select(Switch).where(Switch.id.in_(ids))).all()

    if not switches:
        raise HTTPException(status_code=404, detail="No switches found for given IDs")

    # 检查是否有未找到的 ID
    found_ids = {s.id for s in switches}
    missing = [str(i) for i in ids if i not in found_ids]

    # 添加备份任务
    for switch in switches:
        background_tasks.add_task(download_config, switch.id)

    message = f"Backups scheduled for {len(switches)} switches"
    if missing:
        message += f". Missing IDs: {', '.join(missing)}"

    return {"message": message}


@router.post("/import")
def import_switches(
    session: SessionDep,
    current_user: CurrentUser,
    file: UploadFile = File(...),
) -> Any:
    """
    Import switches from Excel.
    """

    # 保存临时文件
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.write(file.file.read())
    tmp.close()

    wb = load_workbook(tmp.name)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise HTTPException(400, "Excel为空")

    headers = [str(h).strip() for h in rows[0]]

    required_headers = {
        "name",
        "ip",
        "branch",
        "inspector_name",
        "inspector_password",
        "login_type",
        "port",
    }

    if not required_headers.issubset(headers):
        raise HTTPException(400, "Excel列不正确")

    data_rows = rows[1:]

    # ---------- 预加载缓存 ----------
    branches = {b.name: b for b in session.exec(select(Branch)).all()}

    inspectors = {
        (i.name, i.password): i for i in session.exec(select(Inspector)).all()
    }

    login_types = {l.name: l for l in session.exec(select(SwitchLoginType)).all()}

    switches = {s.ip: s for s in session.exec(select(Switch)).all()}

    created = 0
    updated = 0
    skipped = 0

    for row in data_rows:

        row_dict = dict(zip(headers, row))

        name = str(row_dict["name"]).strip()
        ip = str(row_dict["ip"]).strip()
        branch_name = str(row_dict["branch"]).strip()
        inspector_name = str(row_dict["inspector_name"]).strip()
        inspector_password = str(row_dict["inspector_password"]).strip()
        login_type_name = str(row_dict["login_type"]).strip()
        stype = str(row_dict["stype"]).strip()
        port = int(row_dict["port"])

        # ---------- branch ----------
        branch = branches.get(branch_name)
        if not branch:
            branch = Branch(name=branch_name)
            session.add(branch)
            session.flush()
            branches[branch_name] = branch

        # ---------- inspector ----------
        inspector_key = (inspector_name, inspector_password)
        inspector = inspectors.get(inspector_key)

        if not inspector:
            inspector = Inspector(
                name=inspector_name,
                password=inspector_password,
            )
            session.add(inspector)
            session.flush()
            inspectors[inspector_key] = inspector

        # ---------- login_type ----------
        login_type = login_types.get(login_type_name)

        if not login_type:
            login_type = SwitchLoginType(name=login_type_name)
            session.add(login_type)
            session.flush()
            login_types[login_type_name] = login_type

        # ---------- switch ----------
        existing = switches.get(ip)

        if existing:
            # 如果完全相同就跳过
            if (
                existing.ip == ip
                and existing.port == port
                and existing.branch_id == branch.id
                and existing.inspector_id == inspector.id
                and existing.login_type_id == login_type.id
                and existing.stype == stype
            ):
                skipped += 1
                continue

            # 更新
            existing.ip = ip
            existing.port = port
            existing.branch = branch
            existing.inspector = inspector
            existing.login_type = login_type
            existing.stype = stype

            updated += 1

        else:
            switch = Switch(
                name=name,
                ip=ip,
                port=port,
                branch=branch,
                inspector=inspector,
                login_type=login_type,
                stype=stype,
            )

            session.add(switch)
            switches[name] = switch

            created += 1

    session.commit()

    return {
        "message": "导入完成",
        "created": created,
        "updated": updated,
        "skipped": skipped,
    }


@router.post("/config", response_model=SwitchConfigShow)
def get_switch_config(session: SessionDep, current_user: CurrentUser, id: uuid.UUID, config_date: date) -> Any:
    """
    Get a switch's configuration.
    """
    switch = session.get(Switch, id)
    if not switch:
        raise HTTPException(status_code=404, detail="设备不存在")

    if not config_date:
        config_date = date.today()

    switch_config = session.exec(
        select(SwitchConfig).where(
            and_(SwitchConfig.switch_id == id, SwitchConfig.created == config_date)
        )
    ).first()

    rows = session.exec(
        select(SwitchConfig.created)
        .where(SwitchConfig.switch_id == id)
        .order_by(SwitchConfig.created.asc())
    ).all()

    if not switch_config:
        raise HTTPException(status_code=404, detail=f"{config_date}设备配置不存在")

    return SwitchConfigShow(
        content=switch_config.get_config(),
        switch=switch_config.switch and SwitchLite.model_validate(switch_config.switch),
        created=switch_config.created,
        updated=switch_config.updated,
        date_list=[d.isoformat() for d in rows],
    )



@router.get("/export", response_class=FileResponse)
def export_switches(
    session: SessionDep,
    current_user: CurrentUser,
) -> Any:
    """
    Export switches to Excel.
    
    Requires authentication. The current_user parameter triggers automatic
    permission validation through dependency injection.
    """
    # Permission check is handled by CurrentUser dependency
    if not current_user:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    stmt = (
        select(Switch)
        .options(
            selectinload(Switch.branch),
            selectinload(Switch.inspector),
            selectinload(Switch.login_type),
        )
    )

    rows = session.exec(stmt)

    # 创建 Excel
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("switches")

    headers = [
        "name",
        "ip",
        "port",
        "login_type",
        "stype",
        "branch",
        "inspector_name",
        "inspector_password",
    ]

    ws.append(headers)

    # 写入数据（逐行）
    for sw in rows:
        ws.append(
            [
                sw.name,
                sw.ip,
                sw.port,
                sw.login_type.name if sw.login_type else "",
                sw.stype if sw.stype else "",
                sw.branch.name if sw.branch else "",
                sw.inspector.name if sw.inspector else "",
                sw.inspector.password if sw.inspector else "",
            ]
        )

    # 临时文件
    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")

    wb.save(tmp_file.name)

    return FileResponse(
        tmp_file.name,
        filename="switches.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
